import subprocess
import sys
import importlib.util
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

def install_package(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def import_module(name, package):
    try:
        importlib.import_module(name)
        print(f"O pacote {name} está instalado.")
    except ImportError:
        print(f"O pacote {name} não está instalado. Instalando...")
        install_package(package)

def check_install_dependencies():
    import_module("openpyxl", "openpyxl")
    import_module("keyboard", "keyboard")

def clear_excel(file_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename=file_name)
        ws = wb.active
        # Apaga todas as linhas abaixo da linha 1
        ws.delete_rows(2, ws.max_row)
        wb.save(file_name)
        messagebox.showinfo("Sucesso", "Todos os dados abaixo da célula A1 foram apagados.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "O arquivo Excel não foi encontrado.")

def save_to_excel(file_name, professor, turma, barcode):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename=file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Adiciona cabeçalhos se o arquivo não existir
        ws['A1'] = "Nome do Professor"
        ws['B1'] = "Turma"
        ws['C1'] = "Timestamp"
        ws['D1'] = "Código de Barras"

    # Verifica se o código de barras já existe e apaga a linha correspondente
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[3].value == barcode:
            ws.delete_rows(row[0].row, 1)
            found = True
            break

    if not found:
        # Salva o nome do professor e a turma na primeira linha disponível
        row = ws.max_row + 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=1, value=professor)
        ws.cell(row=row, column=2, value=turma)
        ws.cell(row=row, column=3, value=timestamp)
        ws.cell(row=row, column=4, value=barcode)
    else:
        messagebox.showinfo("Informação", f"O código de barras {barcode} foi removido do arquivo Excel.")

    wb.save(file_name)

def main():
    check_install_dependencies()

    file_name = "controlepatrimonio.xlsx"

    def clear_excel_callback():
        if messagebox.askyesno("Confirmar", "Gostaria de apagar todos os dados abaixo da célula A1 no arquivo Excel?"):
            clear_excel(file_name)

    def start_scanning():
        professor = professor_entry.get()
        turma = turma_entry.get()

        if not professor or not turma:
            messagebox.showerror("Erro", "Por favor, preencha os campos Nome do Professor e Turma antes de iniciar a leitura.")
            return

        barcode_label.config(text="Aguardando leitura do código...")

        barcode = ""

        def on_key_event(event):
            nonlocal barcode
            if event.event_type == keyboard.KEY_DOWN:
                if event.name == 'enter':
                    if barcode.strip():  # Verifica se há algo no barcode para evitar strings vazias
                        barcode_label.config(text=f"Código de barras lido: {barcode}")
                        save_to_excel(file_name, professor, turma, barcode)
                        barcode = ""  # Reset barcode after saving
                    else:
                        barcode_label.config(text="Nenhum código lido.")
                elif len(event.name) == 1:  # Verifica se é um caractere simples
                    barcode += event.name

        try:
            import keyboard
            keyboard.hook(on_key_event)
            messagebox.showinfo("Instruções", "Aperte Enter após ler o código de barras.")
            root.mainloop()
        except Exception as e:
            print("Ocorreu um erro:", e)

    root = tk.Tk()
    root.title("Leitor de Códigos de Barras")

    tk.Label(root, text="Nome do Professor:").grid(row=0, column=0, padx=10, pady=10)
    professor_entry = tk.Entry(root)
    professor_entry.grid(row=0, column=1, padx=10, pady=10)

    tk.Label(root, text="Turma:").grid(row=1, column=0, padx=10, pady=10)
    turma_entry = tk.Entry(root)
    turma_entry.grid(row=1, column=1, padx=10, pady=10)

    tk.Button(root, text="Limpar Excel", command=clear_excel_callback).grid(row=2, column=0, padx=10, pady=10)
    tk.Button(root, text="Iniciar Leitura de Códigos", command=start_scanning).grid(row=2, column=1, padx=10, pady=10)

    barcode_label = tk.Label(root, text="")
    barcode_label.grid(row=3, columnspan=2, padx=10, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
